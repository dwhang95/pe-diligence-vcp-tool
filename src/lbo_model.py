"""
lbo_model.py — LBO model Excel generator for the PE Ops Tool Suite.

Generates a 4-sheet openpyxl workbook:
  1. Assumptions   — all inputs (yellow) and derived values (blue)
  2. Income Statement — 5-year P&L linked to Assumptions
  3. Debt Schedule — opening/amort/closing by year, linked to Assumptions
  4. Returns Summary — MOIC + IRR 3x3 sensitivity table

Returns raw bytes (.xlsx). All assumption cells carry a placeholder warning.
"""

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ─────────────────────────────────────────────────────────────────
# Color palette — navy / gold to match the app
# ─────────────────────────────────────────────────────────────────
NAVY        = "1A274A"
MED_NAVY    = "2D3F6B"
GOLD        = "C9A84C"
LIGHT_GOLD  = "FDF3D7"   # input (yellow) cell background
LIGHT_BLUE  = "EDF0F7"   # derived (blue) cell background
WHITE       = "FFFFFF"
BODY        = "1A1A1A"
SUBTLE      = "777777"
WARN        = "B85C00"


# ─────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────

def _fill(c):
    return PatternFill("solid", fgColor=c)

def _font(bold=False, italic=False, color=BODY, size=10, name="Calibri"):
    return Font(bold=bold, italic=italic, color=color, size=size, name=name)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _gold_bottom():
    return Border(bottom=Side(style="medium", color=GOLD))

def _hair_bottom():
    return Border(bottom=Side(style="hair", color="CCCCCC"))

def _col(ws, letter, width):
    ws.column_dimensions[letter].width = width

def _row_h(ws, row, height):
    ws.row_dimensions[row].height = height


# ─────────────────────────────────────────────────────────────────
# Shared layout primitives
# ─────────────────────────────────────────────────────────────────

def _title(ws, row, title_text, subtitle_text):
    c = ws.cell(row=row, column=1, value=title_text)
    c.font = _font(bold=True, size=13, color=NAVY)
    s = ws.cell(row=row + 1, column=1, value=subtitle_text)
    s.font = _font(italic=True, size=8, color=SUBTLE)


def _section_header(ws, row, text, ncols=3):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(bold=True, size=9, color=WHITE)
    c.fill = _fill(MED_NAVY)
    c.alignment = _align("left")
    _row_h(ws, row, 18)


def _input_cell(ws, row, label, value, note="Placeholder — replace with diligence findings",
                fmt=None):
    """Yellow input cell — user should update these."""
    ws.cell(row=row, column=1, value=label).font = _font(size=10)
    v = ws.cell(row=row, column=2, value=value)
    v.font  = _font(bold=True, size=10, color=NAVY)
    v.fill  = _fill(LIGHT_GOLD)
    v.alignment = _align("right")
    v.border = _thin_border()
    if fmt:
        v.number_format = fmt
    n = ws.cell(row=row, column=3, value=note)
    n.font  = _font(italic=True, size=8, color=SUBTLE)
    n.alignment = _align(wrap=True)
    return v


def _derived_cell(ws, row, label, formula, note="Derived", fmt=None):
    """Blue derived / formula cell."""
    ws.cell(row=row, column=1, value=label).font = _font(size=10)
    v = ws.cell(row=row, column=2, value=formula)
    v.font  = _font(size=10, color=BODY)
    v.fill  = _fill(LIGHT_BLUE)
    v.alignment = _align("right")
    v.border = _thin_border()
    if fmt:
        v.number_format = fmt
    n = ws.cell(row=row, column=3, value=note)
    n.font  = _font(italic=True, size=8, color=SUBTLE)
    n.alignment = _align(wrap=True)
    return v


# ─────────────────────────────────────────────────────────────────
# EV range parser
# ─────────────────────────────────────────────────────────────────

def _parse_ev_midpoint(ev_range: str) -> float:
    """'$100–250M' → 175.0"""
    nums = re.findall(r"\d+", ev_range.replace(",", ""))
    if len(nums) >= 2:
        return (float(nums[0]) + float(nums[1])) / 2
    if len(nums) == 1:
        return float(nums[0])
    return 175.0


# ─────────────────────────────────────────────────────────────────
# Sheet 1: Assumptions
# ─────────────────────────────────────────────────────────────────
#
# Row layout (fixed — other sheets reference these by row number):
#
#   1   Title
#   2   Subtitle
#   3   (blank)
#   4   ── DEAL STRUCTURE ──────────────────
#   5   Entry EV ($M)                  B5  INPUT
#   6   LTM EBITDA ($M)                B6  INPUT
#   7   Entry EV/EBITDA Multiple       B7  =B5/B6
#   8   (blank)
#   9   ── CAPITAL STRUCTURE ───────────────
#  10   Senior Debt (%)                B10 INPUT
#  11   Equity (%)                     B11 =1-B10
#  12   Senior Debt ($M)               B12 =B5*B10
#  13   Equity Invested ($M)           B13 =B5*(1-B10)
#  14   Interest Rate — Senior         B14 INPUT
#  15   (blank)
#  16   ── OPERATING ASSUMPTIONS ───────────
#  17   Revenue at Entry ($M)          B17 INPUT
#  18   EBITDA Margin at Entry         B18 =B6/B17
#  19   Revenue Growth Rate (p.a.)     B19 INPUT
#  20   EBITDA Margin (hold period)    B20 =B18
#  21   D&A % of Revenue               B21 INPUT
#  22   Tax Rate                       B22 INPUT
#  23   (blank)
#  24   ── DEBT ASSUMPTIONS ────────────────
#  25   Annual Amortization (% open)   B25 INPUT
#  26   (blank)
#  27   ── EXIT ASSUMPTIONS ────────────────
#  28   Hold Period (years)            B28 INPUT
#  29   Exit Multiple — Low            B29 INPUT
#  30   Exit Multiple — Base           B30 INPUT
#  31   Exit Multiple — High           B31 INPUT

def _build_assumptions(ws, entry_ev: float, industry: str):
    _col(ws, "A", 38)
    _col(ws, "B", 16)
    _col(ws, "C", 54)

    ltm_ebitda  = round(entry_ev / 8.5, 1)
    entry_rev   = round(ltm_ebitda / 0.15, 1)

    _title(ws, 1, "LBO MODEL — ASSUMPTIONS",
           f"{industry}  |  $M unless noted  |  "
           "Yellow = input (replace with diligence findings)  |  Blue = derived")

    # ── Deal Structure ──────────────────────────────────────────
    _section_header(ws, 4, "DEAL STRUCTURE")
    _input_cell(ws,  5, "Entry EV ($M)",            entry_ev,   fmt="#,##0.0")
    _input_cell(ws,  6, "LTM EBITDA ($M)",          ltm_ebitda, fmt="#,##0.0")
    _derived_cell(ws, 7, "Entry EV/EBITDA Multiple", "=B5/B6",
                  note="Derived", fmt='0.0"x"')

    # ── Capital Structure ───────────────────────────────────────
    _section_header(ws, 9, "CAPITAL STRUCTURE")
    _input_cell(ws,  10, "Senior Debt (%)",         0.50,  fmt="0%")
    _derived_cell(ws, 11, "Equity (%)",             "=1-B10",      fmt="0%")
    _derived_cell(ws, 12, "Senior Debt ($M)",       "=B5*B10",     fmt="#,##0.0")
    _derived_cell(ws, 13, "Equity Invested ($M)",   "=B5*(1-B10)", fmt="#,##0.0")
    _input_cell(ws,  14, "Interest Rate — Senior",  0.07,  fmt="0.0%")

    # ── Operating Assumptions ───────────────────────────────────
    _section_header(ws, 16, "OPERATING ASSUMPTIONS")
    _input_cell(ws,  17, "Revenue at Entry ($M)",        entry_rev,  fmt="#,##0.0")
    _derived_cell(ws, 18, "EBITDA Margin at Entry",      "=B6/B17",
                  note="Derived from LTM EBITDA ÷ Entry Revenue", fmt="0.0%")
    _input_cell(ws,  19, "Revenue Growth Rate (p.a.)",   0.05,  fmt="0.0%")
    _derived_cell(ws, 20, "EBITDA Margin (hold period)", "=B18",
                  note="Flat margin assumed — adjust if margin expansion is part of the thesis",
                  fmt="0.0%")
    _input_cell(ws,  21, "D&A % of Revenue",             0.03,  fmt="0.0%")
    _input_cell(ws,  22, "Tax Rate",                     0.26,  fmt="0%")

    # ── Debt Assumptions ────────────────────────────────────────
    _section_header(ws, 24, "DEBT ASSUMPTIONS")
    _input_cell(ws,  25, "Annual Amortization (% of opening balance)", 0.10, fmt="0%",
                note="Placeholder — 10% mandatory cash sweep; update from credit agreement")

    # ── Exit Assumptions ────────────────────────────────────────
    _section_header(ws, 27, "EXIT ASSUMPTIONS")
    _input_cell(ws,  28, "Hold Period (years)",   5,    fmt="0")
    _input_cell(ws,  29, "Exit Multiple — Low",   7.0,  fmt='0.0"x"')
    _input_cell(ws,  30, "Exit Multiple — Base",  8.5,  fmt='0.0"x"')
    _input_cell(ws,  31, "Exit Multiple — High",  10.0, fmt='0.0"x"')

    # Blank row heights
    for blank in [3, 8, 15, 23, 26]:
        _row_h(ws, blank, 8)
    _row_h(ws, 2, 14)


# ─────────────────────────────────────────────────────────────────
# Sheet 2: Income Statement
# ─────────────────────────────────────────────────────────────────
#
# Columns: A=label, B=spacer(4), C=Y1, D=Y2, E=Y3, F=Y4, G=Y5
# Row 4: year headers
# Row 5: Revenue
# Row 6: Revenue Growth %
# Row 7: EBITDA
# Row 8: EBITDA Margin
# Row 9: D&A
# Row 10: EBIT
# Row 11: Interest Expense  ← references Debt Schedule opening balance
# Row 12: EBT
# Row 13: Taxes
# Row 14: Net Income

_YR_LETTERS = ["C", "D", "E", "F", "G"]
_YR_LABELS  = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
_YR_COLS    = [3, 4, 5, 6, 7]


def _build_income_statement(ws):
    _col(ws, "A", 34)
    _col(ws, "B", 4)
    for c in _YR_LETTERS:
        _col(ws, c, 14)

    _title(ws, 1, "INCOME STATEMENT",
           "All cells formula-linked to Assumptions sheet — update inputs there")

    # ── Year headers ──────────────────────────────────────────────
    ws.cell(row=4, column=1, value="($M)").font = _font(italic=True, size=9, color=SUBTLE)
    for yr_label, col in zip(_YR_LABELS, _YR_COLS):
        c = ws.cell(row=4, column=col, value=yr_label)
        c.font      = _font(bold=True, size=10, color=WHITE)
        c.fill      = _fill(NAVY)
        c.alignment = _align("center")
    _row_h(ws, 4, 20)

    def data_row(row, label, formulas, fmt, bold=False, bg=None, total=False):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(bold=bold, size=10, color=MED_NAVY if bold else BODY)
        if bg:
            lc.fill = _fill(bg)
        for col, f in zip(_YR_COLS, formulas):
            vc = ws.cell(row=row, column=col, value=f)
            vc.font      = _font(bold=bold, size=10)
            vc.alignment = _align("right")
            vc.number_format = fmt
            if bg:
                vc.fill = _fill(bg)
            vc.border = _gold_bottom() if total else _hair_bottom()

    L = _YR_LETTERS  # shorthand

    # Revenue: Year 1 = entry * (1+g)^1, Year 2 = entry * (1+g)^2, …
    data_row(5, "Revenue",
             [f"=Assumptions!$B$17*(1+Assumptions!$B$19)^{n}" for n in range(1, 6)],
             "#,##0.0")

    data_row(6, "  Revenue Growth (YoY)",
             ["=Assumptions!$B$19"] * 5,
             "0.0%")

    data_row(7, "EBITDA",
             [f"={c}5*Assumptions!$B$20" for c in L],
             "#,##0.0", bold=True, bg="F0F4FF")

    data_row(8, "  EBITDA Margin",
             [f"={c}7/{c}5" for c in L],
             "0.0%")

    data_row(9, "D&A",
             [f"=-{c}5*Assumptions!$B$21" for c in L],
             "#,##0.0")

    data_row(10, "EBIT",
             [f"={c}7+{c}9" for c in L],
             "#,##0.0", bold=True)

    # Interest expense: opening debt balance from Debt Schedule × interest rate
    data_row(11, "Interest Expense",
             [f"=-'Debt Schedule'!{c}5*Assumptions!$B$14" for c in L],
             "#,##0.0")

    data_row(12, "EBT",
             [f"={c}10+{c}11" for c in L],
             "#,##0.0")

    data_row(13, "Taxes",
             [f"=IF({c}12>0,-{c}12*Assumptions!$B$22,0)" for c in L],
             "#,##0.0")

    data_row(14, "Net Income",
             [f"={c}12+{c}13" for c in L],
             "#,##0.0", bold=True, bg=LIGHT_BLUE, total=True)

    _row_h(ws, 3, 8)


# ─────────────────────────────────────────────────────────────────
# Sheet 3: Debt Schedule
# ─────────────────────────────────────────────────────────────────
#
# Columns: A=label, B=spacer, C=Y1, D=Y2, E=Y3, F=Y4, G=Y5
# Row 4: year headers
# Row 5: Opening Balance  ← C5 = Assumptions!B12; D5-G5 = prior year closing
# Row 6: Annual Amortization
# Row 7: Closing Balance

def _build_debt_schedule(ws):
    _col(ws, "A", 38)
    _col(ws, "B", 4)
    for c in _YR_LETTERS:
        _col(ws, c, 14)

    _title(ws, 1, "DEBT SCHEDULE",
           "Senior debt only — single tranche — opening balance, amortization, closing balance")

    # ── Year headers ──────────────────────────────────────────────
    ws.cell(row=4, column=1, value="($M)").font = _font(italic=True, size=9, color=SUBTLE)
    for yr_label, col in zip(_YR_LABELS, _YR_COLS):
        c = ws.cell(row=4, column=col, value=yr_label)
        c.font      = _font(bold=True, size=10, color=WHITE)
        c.fill      = _fill(NAVY)
        c.alignment = _align("center")
    _row_h(ws, 4, 20)

    def drow(row, label, formulas, fmt="#,##0.0", bold=False, bg=None, total=False):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = _font(bold=bold, size=10, color=MED_NAVY if bold else BODY)
        if bg:
            lc.fill = _fill(bg)
        for col, f in zip(_YR_COLS, formulas):
            vc = ws.cell(row=row, column=col, value=f)
            vc.font      = _font(bold=bold, size=10)
            vc.alignment = _align("right")
            vc.number_format = fmt
            if bg:
                vc.fill = _fill(bg)
            vc.border = _gold_bottom() if total else _hair_bottom()

    # Opening Balance: Y1 from Assumptions!B12; Y2+ from prior closing
    drow(5, "Opening Balance",
         ["=Assumptions!$B$12", "=C7", "=D7", "=E7", "=F7"],
         bold=True, bg="F0F4FF")

    # Amortization: each year = opening × amort rate (declining balance)
    drow(6, "Annual Amortization",
         [f"=-{c}5*Assumptions!$B$25" for c in _YR_LETTERS])

    # Closing Balance
    drow(7, "Closing Balance",
         [f"={c}5+{c}6" for c in _YR_LETTERS],
         bold=True, bg=LIGHT_BLUE, total=True)

    # Footer note
    _row_h(ws, 9, 28)
    note = ws.cell(row=9, column=1,
                   value="Note: Amortization = opening balance × amort rate (declining balance). "
                         "Single-tranche model only — no mezz, no PIK, no revolver. "
                         "Update from credit agreement before use.")
    note.font      = _font(italic=True, size=8, color=SUBTLE)
    note.alignment = _align(wrap=True)
    ws.merge_cells("A9:G9")

    _row_h(ws, 3, 8)


# ─────────────────────────────────────────────────────────────────
# Sheet 4: Returns Summary
# ─────────────────────────────────────────────────────────────────
#
# Columns: A=label(30), B=value/spacer(14), C=3yr(15), D=4yr(15), E=5yr(15)
#
# Rows 4–10:  Key parameters (pulled from Assumptions)
# Rows 12–16: MOIC 3×3 sensitivity (exit mult × hold period)
# Rows 18–22: IRR  3×3 sensitivity
# Row 24:     Disclaimer

def _build_returns_summary(ws):
    _col(ws, "A", 30)
    _col(ws, "B", 14)
    _col(ws, "C", 15)
    _col(ws, "D", 15)
    _col(ws, "E", 15)

    _title(ws, 1, "RETURNS SUMMARY",
           "MOIC and IRR across exit multiples and hold periods — all from Assumptions sheet")

    # ── Key Parameters ─────────────────────────────────────────────
    _section_header(ws, 4, "KEY PARAMETERS  (from Assumptions)", ncols=5)

    def krow(row, label, formula, fmt):
        ws.cell(row=row, column=1, value=label).font = _font(size=10)
        v = ws.cell(row=row, column=2, value=formula)
        v.font      = _font(bold=True, size=10, color=NAVY)
        v.fill      = _fill(LIGHT_BLUE)
        v.alignment = _align("right")
        v.number_format = fmt
        v.border = _thin_border()

    krow(5,  "Entry EV ($M)",          "=Assumptions!B5",  "#,##0.0")
    krow(6,  "Equity Invested ($M)",   "=Assumptions!B13", "#,##0.0")
    krow(7,  "Senior Debt ($M)",       "=Assumptions!B12", "#,##0.0")
    krow(8,  "Revenue at Entry ($M)",  "=Assumptions!B17", "#,##0.0")
    krow(9,  "EBITDA Margin (entry)",  "=Assumptions!B18", "0.0%")
    krow(10, "Revenue Growth (p.a.)",  "=Assumptions!B19", "0.0%")

    _row_h(ws, 11, 8)

    # ── Shared formula builders ────────────────────────────────────
    # MOIC = (exit_EV - debt_at_exit) / equity_invested
    #   exit_EV    = exit_mult × revenue_entry × (1+g)^N × EBITDA_margin_hold
    #   debt_exit  = senior_debt × (1 - amort_rate)^N
    #   equity_inv = Assumptions!B13
    def _moic(mult_cell: str, years: int) -> str:
        return (
            f"=MAX(0,"
            f"(Assumptions!{mult_cell}"
            f"*Assumptions!$B$17*(1+Assumptions!$B$19)^{years}*Assumptions!$B$20"
            f"-Assumptions!$B$12*(1-Assumptions!$B$25)^{years})"
            f"/Assumptions!$B$13)"
        )

    # IRR = MOIC^(1/N) - 1  (simple 2-cash-flow approximation)
    def _irr(mult_cell: str, years: int) -> str:
        return (
            f'=IFERROR(MAX(0,'
            f'(Assumptions!{mult_cell}'
            f'*Assumptions!$B$17*(1+Assumptions!$B$19)^{years}*Assumptions!$B$20'
            f'-Assumptions!$B$12*(1-Assumptions!$B$25)^{years})'
            f'/Assumptions!$B$13)^(1/{years})-1,"N/A")'
        )

    EXIT_ROWS = [
        ("$B$29", "Low"),
        ("$B$30", "Base"),
        ("$B$31", "High"),
    ]
    HOLD_COLS = [(3, 3), (4, 4), (5, 5)]  # (column index, hold years)

    def _sensitivity_table(header_row: int, data_start_row: int,
                           formula_fn, fmt: str, title: str):
        _section_header(ws, header_row,
                        f"{title}  (Exit Multiple ↓  ×  Hold Period →)", ncols=5)

        # Column headers
        ws.cell(row=data_start_row, column=1,
                value="Exit Multiple").font = _font(bold=True, size=9, color=WHITE)
        ws.cell(row=data_start_row, column=1).fill = _fill(NAVY)
        for col, years in HOLD_COLS:
            c = ws.cell(row=data_start_row, column=col, value=f"{years} Years")
            c.font      = _font(bold=True, size=9, color=WHITE)
            c.fill      = _fill(NAVY)
            c.alignment = _align("center")
        _row_h(ws, data_start_row, 18)

        for i, (mult_cell, mult_label) in enumerate(EXIT_ROWS):
            row = data_start_row + 1 + i
            is_base = (mult_label == "Base")

            # Row label — dynamic: reads exit multiple from Assumptions
            lc = ws.cell(row=row, column=1)
            lc.value = f'="{mult_label} ("&TEXT(Assumptions!{mult_cell},"0.0")&"x)"'
            lc.font  = _font(bold=is_base, size=10)
            if is_base:
                lc.fill = _fill(LIGHT_GOLD)

            # Data cells
            for col, years in HOLD_COLS:
                vc = ws.cell(row=row, column=col, value=formula_fn(mult_cell, years))
                vc.font      = _font(bold=is_base, size=10, color=NAVY)
                vc.alignment = _align("center")
                vc.number_format = fmt
                vc.border = _thin_border()
                if is_base:
                    vc.fill = _fill(LIGHT_GOLD)

    _sensitivity_table(12, 13, _moic, '0.0"x"', "MOIC SENSITIVITY")
    _row_h(ws, 17, 8)
    _sensitivity_table(18, 19, _irr,  "0.0%",    "IRR SENSITIVITY")

    # ── Disclaimer ─────────────────────────────────────────────────
    _row_h(ws, 24, 32)
    note = ws.cell(row=24, column=1,
                   value="\u26a0  All figures are model outputs from placeholder assumptions. "
                         "IRR = MOIC^(1/N)\u22121 (simple, no interim cash flows). "
                         "MOIC and IRR will read zero or N/A until assumptions are populated. "
                         "Update the Assumptions sheet with diligence findings before presenting to IC.")
    note.font      = _font(italic=True, size=8, color=WARN)
    note.alignment = _align(wrap=True)
    ws.merge_cells("A24:E24")

    _row_h(ws, 3, 8)
    _row_h(ws, 11, 8)
    _row_h(ws, 17, 8)
    _row_h(ws, 23, 8)


# ─────────────────────────────────────────────────────────────────
# Main entry point
# ─────────────────────────────────────────────────────────────────

def build_lbo_excel(company_name: str, industry: str, ev_range: str) -> bytes:
    """
    Build a 4-sheet LBO model workbook and return raw .xlsx bytes.

    Args:
        company_name: target company (used for context only)
        industry:     industry vertical (shown in Assumptions subtitle)
        ev_range:     EV range string, e.g. '$100–250M' (midpoint used as Entry EV seed)
    """
    entry_ev = _parse_ev_midpoint(ev_range)

    wb = Workbook()

    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a.sheet_view.showGridLines = False
    _build_assumptions(ws_a, entry_ev, industry)

    ws_is = wb.create_sheet("Income Statement")
    ws_is.sheet_view.showGridLines = False
    _build_income_statement(ws_is)

    ws_ds = wb.create_sheet("Debt Schedule")
    ws_ds.sheet_view.showGridLines = False
    _build_debt_schedule(ws_ds)

    ws_r = wb.create_sheet("Returns Summary")
    ws_r.sheet_view.showGridLines = False
    _build_returns_summary(ws_r)

    # Tab accent colors
    ws_a.sheet_properties.tabColor  = NAVY
    ws_is.sheet_properties.tabColor = MED_NAVY
    ws_ds.sheet_properties.tabColor = MED_NAVY
    ws_r.sheet_properties.tabColor  = GOLD

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
